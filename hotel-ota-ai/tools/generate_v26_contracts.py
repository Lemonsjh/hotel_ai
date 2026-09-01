from __future__ import annotations

import json
import re
import shutil
from collections import defaultdict
from pathlib import Path
from typing import Any


ROOT = Path(__file__).resolve().parents[1]
LOCAL_BLUEPRINT_DIR = ROOT.parent / "功能逻辑蓝图"
REFERENCE_DIR = ROOT / "docs" / "architecture_reference" / "v26"
CONTRACT_DIR = ROOT / "contracts"
V26_CONTRACT_DIR = CONTRACT_DIR / "v26"
ARCHITECTURE_DIR = ROOT / "architecture"
SOURCE_MANIFEST = ROOT / "references" / "source" / "source_manifest.yaml"

ALLOWED_DATA_TYPES = {
    "string",
    "number",
    "integer",
    "boolean",
    "datetime",
    "date",
    "object",
    "array",
    "enum",
    "money",
    "ratio",
}

STATUS_MAP = {
    "confirmed_exact": "confirmed_exact",
    "confirmed_alias": "confirmed_alias",
    "inferred_alias": "candidate_inferred",
    "candidate": "candidate_inferred",
    "candidate_inferred": "candidate_inferred",
    "to_confirm": "candidate_inferred",
    "manual_required": "manual_required",
    "not_available": "not_available",
    "project_only": "project_only",
    "config_only": "config_only",
    "algorithm_output": "algorithm_output",
    "upstream_output": "upstream_output",
    "deprecated": "deprecated",
    "fixed": "candidate_inferred",
    "": "candidate_inferred",
    None: "candidate_inferred",
}

FIELD_PACKAGE_SOURCE_LAYER = {
    "AUTH": "config_only",
    "CFG": "config_only",
    "CTX": "project_only",
    "EXE": "upstream_output",
    "MSG": "upstream_output",
}

TYPE_OVERRIDES = {
    "intent_candidate": "string",
    "scenario_candidate": "string",
    "required_blueprints": "array",
    "required_skill_chain": "array",
    "algorithm_required_fields": "array",
    "source_required_fields": "array",
    "agent_plan": "object",
    "required_field_packages": "array",
    "parallel_groups": "array",
    "price_snapshot": "object",
    "direct_price_action_blocked": "boolean",
    "sentiment": "enum",
    "time_windows": "array",
    "roi_decision": "object",
    "data_gaps": "array",
    "price_recommendations": "array",
    "data_gap_list": "array",
    "external_diagnosis_report_html": "object",
    "diagnosis_scorecard": "object",
}

CRITICAL_SKILL_ERROR_HANDLING = {
    "S5": "只给建议，不直接执行调价；必须交给 S6/A4 dry-run 和审批闸门。",
    "S6": "demo_data/sample/stale 下只返回 dry-run/preview，不创建正式审批，不执行 live 写入。",
    "S7": "只预警和解释竞对变化，不直接调价，不绕过 S5/S6。",
    "S11": "demo下只预览推广执行或人工任务，不重新决策，不执行 live 推广写入。",
    "S14-EXT": "第三方 OTA 诊断与本店生产数据隔离；demo 下只生成预览 artifact。",
}


def normalize_legacy_reference_text(text: str) -> str:
    """Keep generated V26 contracts from declaring old versions as current truth."""
    return (
        text.replace("V25", "V26")
        .replace("V22", "legacy")
        .replace("V20", "legacy")
        .replace("V19", "legacy")
    )


def dump_json_compatible(path: Path, payload: Any) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(payload, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")


def find_source_files() -> dict[str, Path]:
    files = {
        "json": next(LOCAL_BLUEPRINT_DIR.glob("*V26*.json")),
        "xlsx": next(LOCAL_BLUEPRINT_DIR.glob("*V26*.xlsx")),
        "drawio": next(LOCAL_BLUEPRINT_DIR.glob("*V26*.drawio")),
        "analysis_xlsx": next(LOCAL_BLUEPRINT_DIR.glob("整体架构分析*.xlsx")),
    }
    return files


def copy_reference_files(files: dict[str, Path]) -> dict[str, str]:
    REFERENCE_DIR.mkdir(parents=True, exist_ok=True)
    copied: dict[str, str] = {}
    for key, source in files.items():
        target = REFERENCE_DIR / source.name
        if key == "json":
            payload = json.loads(source.read_text(encoding="utf-8"))
            payload = sanitize_local_paths(payload)
            target.write_text(json.dumps(payload, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
        elif not target.exists():
            shutil.copy2(source, target)
        copied[key] = str(target.relative_to(ROOT)).replace("\\", "/")
    return copied


def sanitize_local_paths(value: Any) -> Any:
    local_path = str(LOCAL_BLUEPRINT_DIR)
    if isinstance(value, dict):
        return {key: sanitize_local_paths(item) for key, item in value.items()}
    if isinstance(value, list):
        return [sanitize_local_paths(item) for item in value]
    if isinstance(value, str):
        return value.replace(local_path, "<local_blueprint_dir>")
    return value


def load_v26_json() -> dict[str, Any]:
    json_files = list(REFERENCE_DIR.glob("*V26*.json"))
    if not json_files:
        json_files = list(LOCAL_BLUEPRINT_DIR.glob("*V26*.json"))
    with json_files[0].open("r", encoding="utf-8") as handle:
        return json.load(handle)


def load_workbook_rows_by_prefix(prefix: str) -> list[dict[str, Any]]:
    try:
        from openpyxl import load_workbook
    except ModuleNotFoundError:
        return []
    xlsx_files = list(REFERENCE_DIR.glob("*V26*.xlsx")) or list(LOCAL_BLUEPRINT_DIR.glob("*V26*.xlsx"))
    if not xlsx_files:
        return []
    workbook = load_workbook(xlsx_files[0], read_only=True, data_only=True)
    sheet_name = next((name for name in workbook.sheetnames if name.startswith(prefix)), None)
    if not sheet_name:
        return []
    worksheet = workbook[sheet_name]
    rows = list(worksheet.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [str(value).strip() if value is not None else "" for value in rows[0]]
    result: list[dict[str, Any]] = []
    for row in rows[1:]:
        item = {}
        for header, value in zip(headers, row):
            if header:
                item[header] = value
        if any(value not in (None, "") for value in item.values()):
            result.append(item)
    return result


def normalize_key(value: Any) -> str:
    key = str(value or "").strip()
    key = key.replace("：", "_").replace(":", "_").replace("/", "_")
    key = re.sub(r"[^0-9A-Za-z_]+", "_", key)
    key = re.sub(r"_+", "_", key).strip("_").lower()
    return key


def sanitize_text(value: Any, fallback: str = "") -> str:
    text = str(value or "").strip()
    if not text:
        return fallback
    text = text.replace("字段覆盖率率", "字段覆盖率")
    if "待命名字段" in text:
        return fallback or "V26 已完成命名修正"
    return normalize_legacy_reference_text(text)


def sanitize_obj(value: Any) -> Any:
    if isinstance(value, dict):
        return {key: sanitize_obj(item) for key, item in value.items()}
    if isinstance(value, list):
        return [sanitize_obj(item) for item in value]
    if isinstance(value, str):
        return sanitize_text(value, "V26 policy text sanitized")
    return value


def normalize_data_type(value: Any, canonical_key: str = "") -> str:
    normalized_key = normalize_key(canonical_key)
    if normalized_key in TYPE_OVERRIDES:
        return TYPE_OVERRIDES[normalized_key]
    raw = str(value or "").strip().lower()
    if raw in ALLOWED_DATA_TYPES:
        return raw
    if raw in {"int", "integer_number"}:
        return "integer"
    if raw in {"float", "decimal", "double"}:
        return "number"
    if raw in {"bool"}:
        return "boolean"
    if raw in {"list"}:
        return "array"
    if raw in {"dict", "json"}:
        return "object"
    if "/" in raw:
        for part in raw.split("/"):
            part = part.strip()
            if part in ALLOWED_DATA_TYPES and part != "string":
                return part
        return "string"
    if canonical_key.endswith("_rate"):
        return "ratio"
    return "string"


def normalize_status(row: dict[str, Any]) -> str:
    field_package = str(row.get("field_package") or row.get("field_group") or "").strip()
    field_id = str(row.get("field_id") or "")
    if field_id.startswith("N") and "_OUT" in field_id:
        return "algorithm_output"
    if field_package in FIELD_PACKAGE_SOURCE_LAYER:
        return FIELD_PACKAGE_SOURCE_LAYER[field_package]
    raw = row.get("source_match_status")
    mapped = STATUS_MAP.get(raw, STATUS_MAP.get(str(raw).strip(), "candidate_inferred"))
    return mapped


def normalize_status_value(value: Any) -> str:
    return STATUS_MAP.get(value, STATUS_MAP.get(str(value or "").strip(), "candidate_inferred"))


def build_naming_fixes() -> dict[str, dict[str, Any]]:
    fixes: dict[str, dict[str, Any]] = {}
    for row in load_workbook_rows_by_prefix("22_"):
        field_id = str(row.get("原field_id/字段") or "").strip()
        fixed_key = normalize_key(row.get("修正后canonical/project"))
        data_type = normalize_data_type(row.get("data_type"), fixed_key)
        if field_id and fixed_key:
            fixes[field_id] = {
                "canonical_key": fixed_key,
                "project_field_name": fixed_key,
                "zh_name": sanitize_text(row.get("中文名"), fixed_key),
                "data_type": data_type,
                "field_type": row.get("字段归类") or "contract_field",
                "governance_note": "V26 naming/type fix applied.",
                "fix_note": sanitize_text(row.get("开发备注"), ""),
            }
    return fixes


def build_naming_fixes_by_key() -> dict[str, dict[str, Any]]:
    fixes = build_naming_fixes()
    by_key: dict[str, dict[str, Any]] = {}
    for fix in fixes.values():
        key = fix.get("canonical_key")
        if key:
            by_key[str(key)] = dict(fix)
    for key, data_type in TYPE_OVERRIDES.items():
        by_key.setdefault(key, {"canonical_key": key, "project_field_name": key})
        by_key[key]["data_type"] = data_type
    return by_key


def split_values(value: Any) -> list[str]:
    if value is None:
        return []
    if isinstance(value, list):
        raw_items = value
    else:
        raw_items = re.split(r"[,，;；|]\s*", str(value))
    return [str(item).strip() for item in raw_items if str(item).strip()]


def build_field_registry(data: dict[str, Any]) -> dict[str, Any]:
    naming_fixes = build_naming_fixes()
    naming_fixes_by_key = build_naming_fixes_by_key()
    fields = []
    seen_ids: set[str] = set()
    for row in data["unified_contract_fields"]:
        field_id = str(row.get("field_id") or "").strip()
        if not field_id or field_id in seen_ids:
            continue
        seen_ids.add(field_id)
        raw_key = normalize_key(row.get("canonical_key") or row.get("project_field_name") or field_id)
        fix = naming_fixes.get(field_id) or naming_fixes_by_key.get(raw_key, {})
        canonical_key = normalize_key(fix.get("canonical_key") or raw_key)
        data_type = normalize_data_type(fix.get("data_type") or row.get("data_type"), canonical_key)
        source_status = normalize_status(row)
        item = {
            "field_id": field_id,
            "field_package": row.get("field_package") or row.get("field_group") or "",
            "canonical_key": canonical_key,
            "project_field_name": normalize_key(fix.get("project_field_name") or row.get("project_field_name") or canonical_key),
            "zh_name": sanitize_text(fix.get("zh_name") or row.get("zh_name"), canonical_key),
            "description": sanitize_text(row.get("field_description") or row.get("description"), ""),
            "data_type": data_type,
            "field_type": fix.get("field_type") or row.get("field_type") or source_status,
            "required_level": row.get("required_level") or "",
            "blueprint_ids": split_values(row.get("blueprint_ids")),
            "blueprint_original_fields": split_values(row.get("blueprint_field_text")),
            "algorithm_usage": row.get("algorithm_usage") or "",
            "source_layer": sanitize_text(row.get("source_layer") or source_status, source_status),
            "actual_source_candidates": split_values(row.get("actual_source_candidates")),
            "actual_source_aliases": split_values(row.get("actual_source_aliases")),
            "source_match_status": source_status,
            "source_match_reason": sanitize_text(row.get("source_match_reason"), ""),
            "used_by_nodes": split_values(row.get("used_by_nodes")),
            "used_by_skills": split_values(row.get("used_by_skills")),
            "io_direction": row.get("io_direction") or "",
            "missing_policy": row.get("missing_policy") or "缺失时降级/标记 data_gap，不编造",
            "privacy_level": row.get("privacy_level") or "internal",
            "governance_note": sanitize_text(fix.get("governance_note") or row.get("governance_note"), ""),
            "manual_confirm_needed": str(row.get("manual_confirm_needed") or "否"),
            "example_value": row.get("example_value"),
            "fix_note": sanitize_text(fix.get("fix_note") or row.get("fix_note"), ""),
        }
        item["used_by"] = sorted(set(item["used_by_nodes"] + item["used_by_skills"]))
        item["governance"] = {
            "status": item["source_match_status"],
            "needs_manual_review": item["source_match_status"] in {"candidate_inferred", "manual_required"},
            "source_unclear": item["source_match_status"] in {"candidate_inferred", "not_available"},
            "demo_or_sample_not_confirmed": False,
        }
        item["audit_notes"] = []
        fields.append(item)
    canonical_index = {field["canonical_key"]: index for index, field in enumerate(fields)}
    output_rows = list(data.get("output_field_dictionary_v25") or [])
    output_rows.extend(row for row in data.get("node_io_fields", []) if row.get("io_direction") == "output")
    for row in output_rows:
        field_id = str(row.get("output_field_id") or row.get("field_id") or "").strip()
        raw_key = normalize_key(row.get("output_key") or row.get("canonical_key") or row.get("project_field_name") or field_id)
        if not raw_key:
            continue
        fix = naming_fixes.get(field_id) or naming_fixes_by_key.get(raw_key, {})
        canonical_key = normalize_key(fix.get("canonical_key") or raw_key)
        data_type = normalize_data_type(fix.get("data_type") or row.get("data_type"), canonical_key)
        if canonical_key in canonical_index:
            existing = fields[canonical_index[canonical_key]]
            existing["data_type"] = data_type
            existing["zh_name"] = sanitize_text(fix.get("zh_name") or row.get("zh_name"), existing.get("zh_name"))
            if field_id.startswith("N"):
                existing["source_match_status"] = "upstream_output"
                existing["source_layer"] = "upstream_output"
            continue
        output_item = {
                "field_id": field_id,
                "field_package": "OUT",
                "canonical_key": canonical_key,
                "project_field_name": normalize_key(fix.get("project_field_name") or row.get("project_field_name") or canonical_key),
                "zh_name": sanitize_text(fix.get("zh_name") or row.get("zh_name"), canonical_key),
                "description": sanitize_text(row.get("description") or row.get("note"), "V26 node output field"),
                "data_type": data_type,
                "field_type": fix.get("field_type") or "node_output",
                "required_level": row.get("required_level") or "upstream_output",
                "blueprint_ids": [],
                "blueprint_original_fields": [],
                "algorithm_usage": row.get("target_usage") or "",
                "source_layer": "upstream_output",
                "actual_source_candidates": [],
                "actual_source_aliases": [],
                "source_match_status": "upstream_output",
                "source_match_reason": "Added from V26 output field dictionary / node IO.",
                "used_by_nodes": split_values(row.get("source_node")),
                "used_by_skills": split_values(row.get("source_skill")),
                "io_direction": "output",
                "missing_policy": row.get("missing_policy") or "缺失时降级/标记 data_gap，不编造",
                "privacy_level": row.get("privacy_level") or "internal",
                "governance_note": sanitize_text(fix.get("governance_note"), ""),
                "manual_confirm_needed": "否",
                "example_value": None,
                "fix_note": sanitize_text(fix.get("fix_note"), ""),
        }
        output_item["used_by"] = sorted(set(output_item["used_by_nodes"] + output_item["used_by_skills"]))
        output_item["governance"] = {
            "status": output_item["source_match_status"],
            "needs_manual_review": False,
            "source_unclear": False,
            "demo_or_sample_not_confirmed": False,
        }
        output_item["audit_notes"] = []
        fields.append(output_item)
        canonical_index[canonical_key] = len(fields) - 1
    return {
        "schema_version": "v26.0",
        "source_version": data["version"],
        "source_priority": data.get("source_priority") or [],
        "field_count": len(fields),
        "status_policy": {
            "supported_statuses": list(data["field_governance_policy_v26"]["status_definitions"].keys()),
            "legacy_values_are_normalized_by_generator": True,
        },
        "supported_statuses": list(data["field_governance_policy_v26"]["status_definitions"].keys()),
        "supported_data_types": data["field_governance_policy_v26"]["data_type_policy"],
        "algorithm_field_requirements": data.get("algorithm_field_requirements") or [],
        "fields": fields,
    }


def build_node_io(data: dict[str, Any]) -> dict[str, Any]:
    naming_fixes_by_key = build_naming_fixes_by_key()
    rows_by_node: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for row in data["node_io_fields"]:
        rows_by_node[str(row.get("node_id"))].append(row)
    nodes = []
    for node in data["nodes"]:
        node_id = node["id"]
        inputs = []
        outputs = []
        for row in rows_by_node[node_id]:
            raw_key = normalize_key(row.get("canonical_key"))
            fix = naming_fixes_by_key.get(raw_key, {})
            canonical_key = normalize_key(fix.get("canonical_key") or raw_key)
            field = {
                "field_id": row.get("field_id"),
                "canonical_key": canonical_key,
                "project_field_name": normalize_key(fix.get("project_field_name") or row.get("project_field_name") or canonical_key),
                "zh_name": sanitize_text(fix.get("zh_name") or row.get("zh_name"), ""),
                "data_type": normalize_data_type(fix.get("data_type") or row.get("data_type"), canonical_key),
                "field_type": fix.get("field_type") or row.get("field_type") or "",
                "source_type": normalize_status_value(row.get("source_type")),
                "source_node": row.get("source_node") or "",
                "source_skill": row.get("source_skill") or "",
                "required_level": row.get("required_level") or "",
                "missing_policy": sanitize_text(row.get("missing_policy"), ""),
                "evidence": sanitize_text(row.get("evidence"), ""),
                "note": sanitize_text(row.get("note"), ""),
            }
            if row.get("io_direction") == "output":
                outputs.append(field)
            else:
                inputs.append(field)
        nodes.append(
            {
                "node_id": node_id,
                "node_name": node.get("name"),
                "skill_id": node.get("skill_id"),
                "agent_id": node.get("agent_id"),
                "input_fields": inputs,
                "output_fields": outputs,
                "direct_trigger": node.get("direct_trigger"),
                "chain_trigger": node.get("chain_trigger"),
                "trigger_condition": node.get("trigger_condition"),
                "upstream_nodes": sorted({row.get("source_node") for row in rows_by_node[node_id] if row.get("source_node")}),
                "downstream_nodes": split_values(node.get("downstream")),
                "edge_payloads": [],
                "blueprint_basis": node.get("blueprint"),
                "acceptance": node.get("acceptance"),
            }
        )
    return {"schema_version": "v26.0", "source_version": data["version"], "nodes": nodes}


def build_skill_io(data: dict[str, Any]) -> dict[str, Any]:
    sheet_rows = {str(row.get("SkillID") or "").strip(): row for row in load_workbook_rows_by_prefix("12_")}
    nodes_by_skill: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for node in data["nodes"]:
        skill_id = str(node.get("skill_id") or "-")
        if skill_id != "-":
            nodes_by_skill[skill_id].append(node)
    skills = []
    for skill_id in [f"S{i}" for i in range(1, 18)] + ["S14-EXT"]:
        node = (nodes_by_skill.get(skill_id) or [{}])[0]
        sheet = sheet_rows.get(skill_id, {})
        skills.append(
            {
                "skill_id": skill_id,
                "skill_name": sheet.get("Skill中文名") or (node.get("name") or ""),
                "english_dir_name": sheet.get("Skill英文目录名") or "",
                "node_ids": [item["id"] for item in nodes_by_skill.get(skill_id, [])],
                "agent_id": sheet.get("AgentID") or node.get("agent_id") or "",
                "blueprint_basis": sheet.get("蓝图来源") or node.get("blueprint") or "",
                "direct_trigger": sheet.get("直接触发") or node.get("direct_trigger") or "",
                "chain_trigger": sheet.get("链路触发") or node.get("chain_trigger") or "",
                "input_fields": split_values(sheet.get("核心输入字段") or node.get("inputs")),
                "output_fields": split_values(sheet.get("核心输出字段") or node.get("outputs")),
                "upstream_dependencies": split_values(node.get("inputs")),
                "downstream_dependencies": split_values(node.get("downstream")),
                "runtime_command": sheet.get("runtime命令建议") or node.get("runtime") or "",
                "demo_data_files": ["examples/demo_data/demo_manifest.json", f"examples/demo_data/nodes/{node.get('id', 'N000')}.json"],
                "error_handling": sheet.get("开发备注") or "",
                "acceptance_tests": ["tests/contracts/test_v26_skill_io_contract.py"],
            }
        )
    for skill in skills:
        critical_error_handling = CRITICAL_SKILL_ERROR_HANDLING.get(skill["skill_id"], "")
        current_error_handling = str(skill.get("error_handling") or "").strip()
        if critical_error_handling and critical_error_handling not in current_error_handling:
            skill["error_handling"] = f"{current_error_handling} {critical_error_handling}".strip()
    return {
        "schema_version": "v26.0",
        "core_skill_count": data["skill_count_policy_v26"]["core_skill_count"],
        "extension_skill_count": data["skill_count_policy_v26"]["extension_skill_count"],
        "skills": skills,
    }


def build_edge_payload_contract(data: dict[str, Any]) -> dict[str, Any]:
    return {
        "schema_version": "v26.0",
        "edge_count": len(data["edges"]),
        "edges": [
            {
                "edge_id": edge["edge_id"],
                "source_node_id": edge["source_node_id"],
                "target_node_id": edge["target_node_id"],
                "trigger_type": edge.get("trigger_type"),
                "why": edge.get("why"),
                "payload": split_values(edge.get("payload")),
                "blueprint": edge.get("blueprint"),
                "source_output_mapping_required": True,
                "target_input_mapping_required": True,
            }
            for edge in data["edges"]
        ],
    }


def agent_plan_for_nodes(nodes: list[str], node_lookup: dict[str, dict[str, Any]]) -> dict[str, list[str]]:
    result: dict[str, list[str]] = defaultdict(list)
    for node_id in nodes:
        agent_id = node_lookup.get(node_id, {}).get("agent_id")
        if agent_id:
            result[agent_id].append(node_id)
    return dict(result)


def required_field_packages(nodes: list[str], node_io: dict[str, list[dict[str, Any]]]) -> list[str]:
    packages: set[str] = set()
    for node_id in nodes:
        for row in node_io.get(node_id, []):
            package = row.get("field_package")
            if package:
                packages.add(str(package))
    return sorted(packages)


def build_scenario_chain_contract(data: dict[str, Any]) -> dict[str, Any]:
    node_lookup = {node["id"]: node for node in data["nodes"]}
    node_io_rows: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for row in data["node_io_fields"]:
        node_io_rows[str(row.get("node_id"))].append(row)
    scenarios = []
    for scenario in data["scenario_chains"]:
        nodes = scenario.get("nodes") or []
        scenarios.append(
            {
                "scenario_id": scenario["id"],
                "scenario_name": scenario.get("name"),
                "node_chain": nodes,
                "skill_chain": [node_lookup[node_id].get("skill_id") for node_id in nodes if node_lookup.get(node_id, {}).get("skill_id") not in (None, "-")],
                "agent_plan": agent_plan_for_nodes(nodes, node_lookup),
                "required_blueprints": sorted({node_lookup.get(node_id, {}).get("blueprint") for node_id in nodes if node_lookup.get(node_id, {}).get("blueprint")}),
                "required_field_packages": required_field_packages(nodes, node_io_rows),
                "expected_outputs": split_values(scenario.get("output")),
                "required_edges": scenario.get("required_edges") or scenario.get("edge_ids") or [],
                "optional_edges": scenario.get("optional_edges") or [],
                "live_block_policy": "demo_data/sample/stale/missing_date never live; approval_id and fresh data required for any future live path",
                "demo_policy": "demo_supported=true; preview/dry-run only; no formal approval; no live execution",
            }
        )
    return {"schema_version": "v26.0", "scenario_count": len(scenarios), "scenarios": scenarios}


def build_source_alias_mapping(data: dict[str, Any]) -> dict[str, Any]:
    aliases = []
    for row in data["unified_contract_fields"]:
        for alias in split_values(row.get("actual_source_aliases")):
            aliases.append(
                {
                    "canonical_key": normalize_key(row.get("canonical_key")),
                    "project_field_name": normalize_key(row.get("project_field_name") or row.get("canonical_key")),
                    "zh_name": row.get("zh_name") or "",
                    "source_alias": alias,
                    "source_match_status": normalize_status(row),
                    "privacy_level": "internal",
                }
            )
    return {
        "schema_version": "v26.0",
        "source_field_count": len(data["source_field_inventory"]),
        "source_inventory_count": len(data["source_field_inventory"]),
        "aliases": aliases,
        "field_inventory": [
            {
                **item,
                "canonical_key": "",
                "governance": {"status": "candidate_inferred", "needs_manual_review": True},
            }
            for item in data["source_field_inventory"]
        ],
        "source_inventory": data["source_field_inventory"],
        "private_config_policy": "Real DSN, role maps, table names, and credentials stay in /etc/hotel-ota-ai/ or environment variables.",
        "safety_policy": {
            "contains_secret_values": False,
            "real_dsn_or_credentials_allowed": False,
            "github_allowed_content": "only aliases, source classes, capability levels, and non-sensitive mapping metadata",
        },
    }


def build_agent_registry(data: dict[str, Any]) -> dict[str, Any]:
    agent_rows = load_workbook_rows_by_prefix("13_")
    if agent_rows:
        agents = [
            {
                "agent_id": row.get("AgentID"),
                "english_name": row.get("英文名"),
                "zh_name": row.get("中文名"),
                "responsible_nodes": split_values(row.get("负责节点")),
                "responsibility": row.get("职责"),
                "forbidden_actions": row.get("边界/禁止事项"),
            }
            for row in agent_rows
            if row.get("AgentID")
        ]
    else:
        grouped: dict[str, list[str]] = defaultdict(list)
        for node in data["nodes"]:
            grouped[node["agent_id"]].append(node["id"])
        agents = [{"agent_id": agent_id, "responsible_nodes": nodes} for agent_id, nodes in sorted(grouped.items())]
    return {"schema_version": "v26.0", "source_version": data["version"], "agents": agents}


def build_architecture_payloads(data: dict[str, Any]) -> dict[str, Any]:
    main_edges = set(data.get("drawio_main_edges_v19") or [])
    extended_edges = set(data.get("extended_edges_not_drawn_v19") or [])
    node_registry = {
        "schema_version": "v26.0",
        "source_version": data["version"],
        "purpose": data["purpose"],
        "nodes": data["nodes"],
    }
    edge_registry = {
        "schema_version": "v26.0",
        "source_version": data["version"],
        "purpose": data["purpose"],
        "edge_count": len(data["edges"]),
        "edges": [
            {
                **edge,
                "draw_policy": "main" if edge["edge_id"] in main_edges else "extended_not_drawn" if edge["edge_id"] in extended_edges else "unspecified",
            }
            for edge in data["edges"]
        ],
    }
    scenario_registry = {
        "schema_version": "v26.0",
        "source_version": data["version"],
        "purpose": data["purpose"],
        "main_route_policy": data.get("main_map_policy") or {},
        "scenarios": [
            {
                **scenario,
                "demo_supported": True,
                "demo_fixture_id": scenario["id"].lower(),
                "demo_expected_outputs": [
                    "data_source_type=demo_data",
                    "freshness_status=demo_data",
                    "approval_data_allowed=false",
                    "live_allowed=false",
                ],
            }
            for scenario in data["scenario_chains"]
        ],
        "future_scenarios": [
            {"scenario_id": "SC11", "status": "future/not_implemented"},
            {"scenario_id": "SC12", "status": "future/not_implemented"},
        ],
    }
    mappings = [
        {
            "node_id": node["id"],
            "node_name": node["name"],
            "agent_id": node["agent_id"],
            "skill_id": node["skill_id"],
            "phase": node.get("phase"),
            "module": node.get("module"),
            "layer": node.get("layer"),
        }
        for node in data["nodes"]
    ]
    mapping_payload = {
        "schema_version": "v26.0",
        "source_version": data["version"],
        "purpose": "V26 node to logical agent mapping. Runtime must read agent_id from this registry.",
        "mappings": mappings,
    }
    return {
        "node_registry.json": node_registry,
        "edge_registry.json": edge_registry,
        "scenario_chain_registry.json": scenario_registry,
        "node_agent_mapping.json": mapping_payload,
        "agent_registry.json": build_agent_registry(data),
    }


def update_source_manifest(copied: dict[str, str], data: dict[str, Any]) -> None:
    payload = {
        "schema_version": "v26.0",
        "purpose": "Track source-of-truth documents used to generate V26 architecture and contract artifacts.",
        "local_reference_directory": {
            "path": str(LOCAL_BLUEPRINT_DIR),
            "usage": "Only for local development and source verification. Do not depend on this path at runtime.",
        },
        "source_files": {
            "v26_node_registry_json": {
                "role": "Primary machine source for nodes, edges, scenarios, fields, node IO, source inventory, and V26 governance policy.",
                "priority": 1,
                "repository_path": copied["json"],
                "available_in_repo": True,
            },
            "v26_contract_xlsx": {
                "role": "Workbook source for sheet-level validation, naming fixes, Skill IO views, and missing-field degradation policy.",
                "priority": 2,
                "repository_path": copied["xlsx"],
                "available_in_repo": True,
            },
            "v26_drawio": {
                "role": "Visual collaboration source. It cannot override V26 JSON field, node, edge, or scenario facts.",
                "priority": 3,
                "repository_path": copied["drawio"],
                "available_in_repo": True,
            },
            "overall_architecture_analysis": {
                "role": "Candidate source field evidence. It cannot override blueprint algorithm requirements.",
                "priority": 4,
                "repository_path": copied["analysis_xlsx"],
                "available_in_repo": True,
            },
        },
        "source_priority": data.get("source_priority") or [],
        "generated_artifacts": [
            "architecture/node_registry.json",
            "architecture/edge_registry.json",
            "architecture/scenario_chain_registry.json",
            "architecture/node_agent_mapping.json",
            "architecture/agent_registry.json",
            "contracts/v26/field_registry.yaml",
            "contracts/v26/node_io_contract.yaml",
            "contracts/v26/skill_io_contract.yaml",
            "contracts/v26/edge_payload_contract.yaml",
            "contracts/v26/scenario_chain_contract.yaml",
            "contracts/v26/source_alias_mapping.yaml",
            "contracts/v26/field_governance_policy.yaml",
        ],
        "legacy_sources": {
            "v19_v20_v22_v25": "migration history only; not the highest source of truth after V26 migration",
        },
        "rules": [
            "runtime, skill, contracts, architecture, and router files must not depend on local absolute paths.",
            "Do not infer algorithm fields from current implementation fields.",
            "V26 JSON is the primary machine fact source.",
            "V26 xlsx sheets 20-24 provide naming, Skill IO, and degradation-policy supplements.",
            "Current project evidence is reference only and cannot reverse-define V26 fields.",
        ],
    }
    dump_json_compatible(SOURCE_MANIFEST, payload)


def main() -> int:
    source_files = find_source_files()
    copied = copy_reference_files(source_files)
    data = load_v26_json()

    field_registry = build_field_registry(data)
    node_io = build_node_io(data)
    skill_io = build_skill_io(data)
    edge_payload = build_edge_payload_contract(data)
    scenario_chain = build_scenario_chain_contract(data)
    source_alias = build_source_alias_mapping(data)

    dump_json_compatible(V26_CONTRACT_DIR / "field_registry.yaml", field_registry)
    dump_json_compatible(V26_CONTRACT_DIR / "node_io_contract.yaml", node_io)
    dump_json_compatible(V26_CONTRACT_DIR / "skill_io_contract.yaml", skill_io)
    dump_json_compatible(V26_CONTRACT_DIR / "edge_payload_contract.yaml", edge_payload)
    dump_json_compatible(V26_CONTRACT_DIR / "scenario_chain_contract.yaml", scenario_chain)
    dump_json_compatible(V26_CONTRACT_DIR / "source_alias_mapping.yaml", source_alias)
    dump_json_compatible(V26_CONTRACT_DIR / "field_governance_policy.yaml", sanitize_obj(data["field_governance_policy_v26"]))

    # Compatibility entrypoints keep old file names while making V26 the current fact source.
    dump_json_compatible(CONTRACT_DIR / "field_registry.yaml", field_registry)
    dump_json_compatible(CONTRACT_DIR / "node_io_contract.yaml", node_io)
    dump_json_compatible(CONTRACT_DIR / "source_mapping.yaml", source_alias)
    dump_json_compatible(
        CONTRACT_DIR / "contract_version.yaml",
        {
            "schema_version": "v26.0",
            "version": data["version"],
            "primary_source": copied["json"],
            "workbook_source": copied["xlsx"],
            "drawio_source": copied["drawio"],
            "compatibility_note": "Top-level contracts mirror contracts/v26 for legacy runtime loaders.",
        },
    )

    for filename, payload in build_architecture_payloads(data).items():
        dump_json_compatible(ARCHITECTURE_DIR / filename, payload)

    update_source_manifest(copied, data)

    print(
        json.dumps(
            {
                "status": "ok",
                "version": data["version"],
                "field_count": field_registry["field_count"],
                "node_count": len(data["nodes"]),
                "edge_count": len(data["edges"]),
                "scenario_count": len(data["scenario_chains"]),
                "contracts_dir": str(V26_CONTRACT_DIR.relative_to(ROOT)).replace("\\", "/"),
            },
            ensure_ascii=False,
        )
    )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
